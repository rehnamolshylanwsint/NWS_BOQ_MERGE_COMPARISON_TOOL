import streamlit as st
import pandas as pd
import re, io
import plotly.express as px
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ================= CONFIG =================
HEADER_SEARCH_ROWS = 30
RATE_PATTERN   = re.compile(r'(RATE)', re.IGNORECASE)
AMOUNT_PATTERN = re.compile(r'(AMOUNT)', re.IGNORECASE)
QTY_PATTERN    = re.compile(r'(QTY|QUANTITY)', re.IGNORECASE)

# Colors
GREEN_FILL  = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")
RED_FILL    = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
HEADER_FILL = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")

# Borders
thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# ================= HELPERS =================
def detect_header_row(df0: pd.DataFrame):
    best_row, best_score = None, -1
    for r in range(min(HEADER_SEARCH_ROWS, len(df0))):
        row_vals = df0.iloc[r].astype(str).str.strip().str.upper().tolist()
        score, has_rate = 0, False
        for v in row_vals:
            if "ITEM" in v: score += 1
            if "DESC" in v or "DESCRIPTION" in v: score += 1
            if "RATE" in v: score, has_rate = score + 3, True
            if "UNIT" in v: score += 1
            if "QUANTITY" in v or "QTY" in v: score += 1
        if has_rate and score > best_score:
            best_score, best_row = score, r
    return best_row

def to_number(x):
    if x is None: return None
    try:
        if isinstance(x, str):
            s = re.sub(r'[^\d\.\-]', '', x.strip())
            if not s: return None
            return float(s)
        return float(x)
    except:
        return None

def shorten_name(name):
    base = name.split('.')[0]
    return re.sub(r'[^A-Za-z0-9]', '', base)[:8]

def clean_unnamed_columns(df, prefix):
    new_cols = []
    for i, col in enumerate(df.columns):
        if str(col).startswith("Unnamed:"):
            new_cols.append(f"{prefix}Col{i+1}")
        else:
            new_cols.append(col)
    df.columns = new_cols
    return df

def style_worksheet(ws, header_row):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if cell.row == header_row:
                cell.font = Font(bold=True, color="000000")
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        max_length = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
            try:
                length = len(str(row[0]))
                if length > max_length: max_length = length
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

# ================= APP STATE =================
st.set_page_config(page_title="Tender BoQ Comparison", layout="wide")

if "page" not in st.session_state:
    st.session_state["page"] = "login"
if "visited" not in st.session_state:
    st.session_state["visited"] = False

# ================= LOGIN PAGE =================
if st.session_state["page"] == "login":
    st.markdown(
        """
        <h1 style="text-align:center; color:#1E3D59;">🔐 Secure Login</h1>
        <p style="text-align:center; color:#444;">Please enter your company password to continue</p>
        """,
        unsafe_allow_html=True
    )
    password = st.text_input("Enter Password:", type="password")
    if st.button("Login"):
        if password == "nws123":  # <<< Set your company password here
            st.session_state["page"] = "welcome"
            st.rerun()
        else:
            st.error("❌ Invalid password. Try again.")

# ================= WELCOME PAGE =================
elif st.session_state["page"] == "welcome":
    st.image("logo.png", width=200)

    if st.session_state["visited"]:
        st.success("👋 Welcome back! Good to see you again.")

    st.markdown(
        """
        <div style="background-color:#f4faff; padding:30px; border-radius:15px;
        box-shadow:0px 4px 12px rgba(0,0,0,0.1); text-align:center;">
        <h1 style="color:#1E3D59;">NWS International</h1>
        <h3>AI-Powered Tender BoQ Merge & Comparison Tool</h3>
        <p style="font-size:16px; color:#333;">
        Upload multiple contractor BoQs, merge them, compare <b>RATES & AMOUNTS</b>,<br>
        and download a professional Excel with lowest/highest/missing highlighted.<br><br>
        <b>Smart. Fast. Professional. 🚀</b>
        </p>
        </div>
        """,
        unsafe_allow_html=True
    )

    if st.button("🚀 Get Started"):
        st.session_state["page"] = "main"
        st.session_state["visited"] = True
        st.rerun()

# ================= MAIN APP =================
elif st.session_state["page"] == "main":
    st.sidebar.header("⚙️ Settings")
    if st.sidebar.button("🏠 Back to Home"):
        st.session_state["page"] = "welcome"
        st.rerun()

    take_first_three_only = st.sidebar.checkbox("Compare only first 3 contractor columns", value=False)

    uploaded_files = st.file_uploader("📂 Upload contractor Excel files", type=["xlsx"], accept_multiple_files=True)

    if uploaded_files and st.button("🔗 Merge & Compare"):
        all_sheets = {}
        for uf in uploaded_files:
            xls = pd.ExcelFile(uf)
            for sheet in xls.sheet_names:
                all_sheets.setdefault(sheet, []).append(uf)

        merged_book = {}
        for sheet_name, files in all_sheets.items():
            dfs = []
            for idx, uf in enumerate(files):
                df = pd.read_excel(uf, sheet_name=sheet_name)
                prefix = shorten_name(uf.name)
                df = clean_unnamed_columns(df, prefix)
                if idx == 0:
                    dfs.append(df)
                else:
                    df = df.iloc[:, 2:]
                    df = df.add_prefix(prefix + "_")
                    dfs.append(df)
            merged_book[sheet_name] = pd.concat(dfs, axis=1)

        bio = io.BytesIO()
        with pd.ExcelWriter(bio, engine="openpyxl") as writer:
            for sheet, df in merged_book.items():
                df.to_excel(writer, sheet_name=sheet, index=False)
        bio.seek(0)
        merged_data = bio.getvalue()

        wb = load_workbook(io.BytesIO(merged_data))
        summary = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            df0 = pd.read_excel(io.BytesIO(merged_data), sheet_name=sheet_name, header=None, dtype=object)
            header_row = detect_header_row(df0)
            if header_row is None:
                continue
            header_excel_row = header_row + 1

            headers = [str(ws.cell(row=header_excel_row, column=c).value or "").strip()
                       for c in range(1, ws.max_column + 1)]

            rate_cols   = [i for i, h in enumerate(headers, 1) if RATE_PATTERN.search(h)]
            amount_cols = [i for i, h in enumerate(headers, 1) if AMOUNT_PATTERN.search(h)]
            qty_col     = next((i for i, h in enumerate(headers, 1) if QTY_PATTERN.search(h)), None)

            if not rate_cols and not amount_cols:
                continue
            if take_first_three_only and len(rate_cols) > 3:
                rate_cols = rate_cols[:3]
                amount_cols = amount_cols[:3]

            low_count, high_count, missing_count = 0, 0, 0
            for r in range(header_excel_row + 1, ws.max_row + 1):
                qty = to_number(ws.cell(row=r, column=qty_col).value) if qty_col else None
                contractor_data = []

                for j, col in enumerate(rate_cols):
                    rate_val = to_number(ws.cell(row=r, column=col).value)
                    amount_val = None
                    if j < len(amount_cols):
                        amount_val = to_number(ws.cell(row=r, column=amount_cols[j]).value)

                    eff_val, rate_valid = None, False
                    if rate_val is not None:
                        eff_val, rate_valid = rate_val, True
                    elif amount_val is not None:
                        eff_val = amount_val / qty if qty not in (None, 0) else amount_val

                    contractor_data.append((col, amount_cols[j] if j < len(amount_cols) else None, eff_val, rate_valid))

                present = [d[2] for d in contractor_data if d[2] is not None]
                if not present:
                    continue

                mn, mx = min(present), max(present)

                for (rate_col, amt_col, eff_val, rate_valid) in contractor_data:
                    rate_cell = ws.cell(row=r, column=rate_col)
                    amt_cell = ws.cell(row=r, column=amt_col) if amt_col else None

                    if eff_val is None:
                        rate_cell.fill = YELLOW_FILL
                        if amt_cell: amt_cell.fill = YELLOW_FILL
                        missing_count += 1
                    else:
                        if not rate_valid:
                            rate_cell.fill = YELLOW_FILL
                        if eff_val == mn:
                            if amt_cell: amt_cell.fill = GREEN_FILL
                            if rate_valid: rate_cell.fill = GREEN_FILL
                            low_count += 1
                        elif eff_val == mx:
                            if amt_cell: amt_cell.fill = RED_FILL
                            if rate_valid: rate_cell.fill = RED_FILL
                            high_count += 1

            style_worksheet(ws, header_excel_row)
            ws.freeze_panes = ws[f"A{header_excel_row+1}"]
            summary.append([sheet_name, low_count, high_count, missing_count])

        out_io = io.BytesIO()
        wb.save(out_io)
        out_io.seek(0)

        # Sidebar summary
        st.sidebar.subheader("📊 Summary")
        df_summary = pd.DataFrame(summary, columns=["Sheet", "Lowest (Green)", "Highest (Red)", "Missing (Yellow)"])
        st.sidebar.dataframe(df_summary, width="stretch")

        # Bubble chart
        if not df_summary.empty:
            df_long = df_summary.melt(id_vars="Sheet", var_name="Type", value_name="Count")
            fig = px.scatter(
                df_long,
                x="Sheet",
                y="Count",
                size="Count",
                color="Type",
                text="Count",
                color_discrete_map={
                    "Lowest (Green)": "#00CC44",
                    "Highest (Red)": "#CC0000",
                    "Missing (Yellow)": "#FFD700",
                },
            )
            fig.update_traces(textposition="top center", marker=dict(line=dict(width=1, color="black"), opacity=0.8))
            fig.update_layout(title="📊 Rate & Amount Comparison Summary", height=500)
            st.plotly_chart(fig, use_container_width=True)

        # Preview
        st.success("✅ Merge & Comparison complete")
        for sheet, df in merged_book.items():
            st.subheader(f"📑 {sheet}")
            st.dataframe(df.head(50), width="stretch")

        st.download_button(
            "⬇️ Download Highlighted Excel",
            out_io,
            file_name="Tender_BoQ_Comparison_Formatted.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# ================= AI ASSISTANT (ALWAYS IN SIDEBAR) =================
st.sidebar.subheader("🤖 AI Assistant")
user_q = st.sidebar.text_input("Ask me anything about BoQ:")

if user_q:
    if "merge" in user_q.lower():
        st.sidebar.info("This tool merges multiple contractor BoQs sheet by sheet.")
    elif "color" in user_q.lower():
        st.sidebar.info("Green = Lowest, Red = Highest, Yellow = Missing values.")
    elif "excel" in user_q.lower():
        st.sidebar.info("You can download a formatted Excel with highlights.")
    else:
        st.sidebar.info("I’m here to help! Try asking about merge, color, or excel.")
